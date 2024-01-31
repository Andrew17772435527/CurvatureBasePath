import numpy as np
import h5py
import matplotlib.pyplot as plt
import math
import copy
import EstimatePath as es
import filter
from openpyxl import load_workbook
from scipy.spatial import KDTree


def Load_Excel(xlsxFileName):
    odometry_data = {'timestamp':[],'hostVel_mps':[],'yawRate_radps':[],'steeringAngle_radps':[],'accel_mps2x':[],'Curvature_radps':[],'CurvatureRate_radpss':[]}
    wb = load_workbook(xlsxFileName)
    ws = wb['Sheet1']
    odometry_data['timestamp'] = [cell.value for cell in ws['A']]
    odometry_data['hostVel_mps'] = [cell.value for cell in ws['B']] 
    odometry_data['yawRate_radps'] = [cell.value for cell in ws['C']] 
    odometry_data['steeringAngle_radps'] = [cell.value for cell in ws['D']] 
    odometry_data['accel_mps2x'] = [cell.value for cell in ws['E']]
    return odometry_data

def Load_Hdf5(h5name):
    odometry_data = {'timestamp':[],'hostVel_mps':[],'yawRate_radps':[],'steeringAngle_radps':[],'accel_mps2x':[],'Curvature_radps':[],'CurvatureRate_radpss':[]}
    file = h5py.File(h5name, 'r')
    dataset = file['odometry']
    odometry_data['timestamp'] = np.zeros(len(dataset))
    odometry_data['hostVel_mps'] = np.zeros(len(dataset))
    odometry_data['yawRate_radps'] = np.zeros(len(dataset))
    odometry_data['steeringAngle_radps'] = np.zeros(len(dataset))
    odometry_data['accel_mps2x'] = np.zeros(len(dataset))
    n = 0
    for frame in dataset:
        odometry_data['timestamp'][n] = frame[0]
        odometry_data['hostVel_mps'][n] = frame[1]
        odometry_data['yawRate_radps'][n] = frame[2]
        odometry_data['steeringAngle_radps'][n] = frame[3]
        odometry_data['accel_mps2x'][n] = frame[4]
        n+=1
    file.close()
    return odometry_data
    
def Limit(input,max,min):
    input_t=copy.copy(input)
    if input>max:
        input_t = max
    elif input<min:
        input_t = min
    return input_t

def SefaDivide(num,den):
    if den != 0:
        return (num/den)
    else:
        return (math.inf)
    

def calculate_path(odometry_data, timestamp):
    l = len(odometry_data["timestamp"])
    pos = np.zeros((2, l))
    heading = 0
    ref_heading = 0
    ref_pos = np.zeros(2)
    for i in range(1, l):
        v = odometry_data["hostVel_mps"][i]
        yr = odometry_data["yawRate_radps"][i]
        delta_t = (odometry_data["timestamp"][i] - odometry_data["timestamp"][i-1]) / 1000
        heading += yr * delta_t
        d = v * delta_t * np.array([np.cos(heading), np.sin(heading)])
        pos[:, i] = pos[:, i - 1] + d

        if timestamp == odometry_data["timestamp"][i]:
            ref_heading = heading
            ref_pos = pos[:, i]

    R = np.array([[np.cos(ref_heading), -np.sin(ref_heading)],
                    [np.sin(ref_heading), np.cos(ref_heading)]])
    pos = R.T @ (pos.T - ref_pos.T).T

    return pos


def plot_od_path(odometry_data, timestamp,n,case):
    """ function to plot odometry data onto the xy plot
    Args:
    odometry_data(np.ndarray) : odometry data
    Returns:
    """
    od_path_data = calculate_path(odometry_data, timestamp)
    if case:
        x = od_path_data[0,:]
        y = od_path_data[1,:]
        end1 = n + 100
        end2 = n + 80
        begin = n-20
        if end1 > len(x)-1:
            end1 = len(x)-1
        if end2 > len(x)-1:
            end2 = len(x)-1
        if begin < 0:
            begin =0
        plt.scatter(x[n:end2],y[n:end2],color='b',s=30)
        plt.plot(x[begin:end1],y[begin:end1],color='k')
    return od_path_data

def find_nearest_point(A, curve):
    kd_tree = KDTree(curve)
    _, nearest_index = kd_tree.query(A)
    nearest_point = curve[nearest_index]
    return nearest_index

if __name__ == '__main__':
    print("-------debug-------")
    VechicleMotionT1 = dict(zip(es.keys, es.values))
    VehSelf = {'ALgt':-1,'VLgt':3}
    odometry_data = {'timestamp':[],'hostVel_mps':[],'yawRate_radps':[],'steeringAngle_radps':[],'accel_mps2x':[],'Curvature_radps':[],'CurvatureRate_radpss':[]}
    h5FileName = "data/Ufront_right_radar_14_30_20_to_14_30_50_1701757820186.h5"
    CorretTracker  = np.array([5,0])
    Path = es.PreDictPath()
    n = 0
    # calculate curvature
    filterWeight  = np.array([0.1,0.1,0.1,0.3,0.4]) # for curvature rate
    filterWeight1 = np.array([0.3,0.1,0.1,0.53,-0.03]) # for lowPassFilter
    odometry_data = Load_Hdf5(h5FileName)
    filter.calculateCurvature(odometry_data)
    filter.LimitChange(odometry_data['Curvature_radps'],0.01)
    filter.calculateCurvatureRate(odometry_data,filterWeight)
    od_path_data = plot_od_path(odometry_data, odometry_data['timestamp'][n],50,False)
    ##
    VehSelf['VLgt'] = odometry_data['hostVel_mps'][n]
    VehSelf['ALgt'] = odometry_data['accel_mps2x'][n]
    Curvature = odometry_data['Curvature_radps'][n]
    CurvatureRate = odometry_data['CurvatureRate_radpss'][n]
    ##
    X,Y = Path.BezierCurvePath(VechicleMotionT1,VehSelf,Curvature,CurvatureRate)
    lenth = np.zeros(len(X))
    curve = np.zeros((len(X),2))
    for i in range(len(X)):
        curve[i,:] = [X[i],Y[i]]
        if i < (len(X)-1):
            lenth[i+1] = lenth[i] + np.sqrt( (X[i+1]-X[i])*(X[i+1]-X[i]) + (Y[i+1]-Y[i])*(Y[i+1]-Y[i]) )
    index = find_nearest_point(CorretTracker, curve)
    plt.scatter(X[index],Y[index],s = 50, color ='g')
    plt.scatter(CorretTracker[0],CorretTracker[1], color ='r')
    lag_L = np.sqrt( (CorretTracker[0]-X[index])*(CorretTracker[0]-X[index]) + (CorretTracker[1]-Y[index])*(CorretTracker[1]-Y[index]) )
    log_S = lenth[index]
    print("SLFrame: ",log_S,",",lag_L)
    D_theta = math.atan(lag_L/log_S)
    print("D_theta: ",D_theta)
    # Path.BezierCurvePath(VechicleMotionT1,VehSelf,Curvature-D_theta,CurvatureRate)
    plt.show()